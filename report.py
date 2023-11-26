import datetime

from kivymd.toast import toast

import config
import os
from openpyxl import Workbook, load_workbook

import inventoryapp


class Report:
    def __init__(self, file, report, lib, param, date=None):
        self.file = self.make_file_result(file, report, lib, param, date)


    @staticmethod
    def make_file_result(file, report, lib, param, date=None):
        if not date:
            date = datetime.date.today().strftime('%Y-%m-%d')
        if param > 5:
            param = ''
            format = 'txt'
        else:
            param = config.PARAMETERS[param - 1]
            format = 'xlsx'
        name = f'/{date}_{lib}_{report}_{param}.{format}'
        if len(file) > 2:
            return file + name
        else:
            path = '\\'.join([os.getcwd(), name])
            file = f'\\{name}'
            try:
                os.mkdir(path)
            except OSError:
                return path + file
            return path + file


    def report_library(self, report:list):
        toast('Запись файлов')

        if os.access(self.file, os.F_OK):
            wb = load_workbook(self.file)
        else:
            wb = Workbook()
        ws = wb.active
        for row in report:
            ws.append(list(row))
        wb.save(self.file)
        name = self.file.split("/")[-1]
        toast(f'Сохранено в файл: {name}')




